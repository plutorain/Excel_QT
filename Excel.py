import sys
import openpyxl
import os
from PyQt5.QtWidgets import *
from PyQt5 import uic

import time

form_class = uic.loadUiType("Excel.ui")[0]

MAX_COL_SIZE = 10
MAX_ROW_SIZE = 10

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

class Exlcol():
    def __init__(self, num):
        self.index = num
        self.index_txt = []
        self.indexlist_up = []
        self.length = 0
        self.coeffcnt = 0
        
        
        
    def int_to_text2(self):
        self.length_check(self.index) #update length info
        self.index_txt = []
        if(self.index == 0):
            print "ERROR index is 0"
            return -1;
        
        coeff_3 = ( self.index/(26**2) ) 
        print(coeff_3)
        coeff_2 = ( self.index - (coeff_3*(26**2)) ) / (26**1)
        print(coeff_2)
        coeff_1 = ( self.index - (coeff_3*(26**2)) - (coeff_2*(26**1)) ) / (26**0) 
        print(coeff_1)
 
        if(coeff_1 == 0): 
            if(coeff_2 == 0):
                #coeff_3 not zero !!!
                coeff_3 = coeff_3 -1
                coeff_2 = 25
                coeff_1 = 26
                
            else:
                coeff_1 = 26
                coeff_2 = coeff_2-1
                if(coeff_2 == 0):
                    if(coeff_3 != 0):
                        coeff_3 = coeff_3 -1
                        coeff_2 = 26
                    else:
                        pass
                else:
                    pass
        
        else:
            if(coeff_2 == 0): 
                if(coeff_3 ==0): # X 0 0
                    pass
                else: # X 0 X
                    coeff_2 = 26
                    coeff_3 = coeff_3-1
            else: # X
                pass
        
        
        if(coeff_3!=0):
            self.index_txt.append(chr(ord('A')-1 + coeff_3))
        
        if(coeff_2!=0):
            self.index_txt.append(chr(ord('A')-1 + coeff_2))
        
        if(coeff_1!=0):
            self.index_txt.append(chr(ord('A')-1 + coeff_1))
        
        self.index_txt = "".join(self.index_txt)
        
        print(self.index_txt)
    
    def int_to_text(self , num):
        self.length_check(num) #update length info
        self.index_txt = []
        if(num == 0):
            print "ERROR index is 0"
            return -1;
        
        coeff = []
        for i in range(0, self.coeffcnt): #make list buffer
            coeff.append("")
        
        reverse = range(self.coeffcnt-1, -1, -1) # n, n-1, ... , 2,1,0 
        
        for i in reverse: # n, n-1, ... , 2,1,0 
            minus_term = 0 
            for j in list(reversed(range(i, self.coeffcnt-1))):
                minus_term = minus_term+(coeff[j+1]*(26**(j+1)))            
            result = (num - minus_term) / (26**i)
            coeff[i] = result
 
        for i in range (0,self.coeffcnt-1):
            if(coeff[i] == 0):
                down_index = 0
                for j in range (i+1, self.coeffcnt): #next to final digit
                    if(coeff[j] > 0):
                        down_index = j
                        break
                coeff[down_index] = coeff[down_index] - 1
                
                for j in range (down_index-1,-1,-1): #digit distance down from existing number digit
                    if(j==i):
                        coeff[j] = 26
                        break
                    else:
                        coeff[j] = 25
                        

        # print(coeff)
        
        for i in range(self.length-1,-1,-1):
            self.index_txt.append(chr(ord('A')-1 + coeff[i]))

        self.index_txt = "".join(self.index_txt)
        # print(self.index_txt)
        return self.index_txt
    
    def list_up(self):
        for i in range(1, self.index+1):
            self.indexlist_up.append(self.int_to_text(i))
        
        return self.indexlist_up
    
    def length_check(self, num):
        temp_value = 26
        temp_remain = 0
        self.length = 1;
        self.coeffcnt = 0
        while not(num <= temp_value):
            # print("temp_value %s" % temp_value)
            self.length = self.length+1
            temp_value = temp_value*( 1 + temp_value )
        
        self.coeffcnt = self.length
        if(num == temp_value):
            self.coeffcnt = self.coeffcnt+1
        
       # print("length : %s coeff_cnt : %s" % (self.length, self.coeffcnt))

class MyWindow(QMainWindow, form_class):
    def __init__(self):
        super(QMainWindow, self).__init__()
        self.setupUi(self)
        self.pushLoadButton.clicked.connect(self.Load_btn_clicked)
        self.pushSaveButton.clicked.connect(self.Save_btn_clicked)
        self.pushTestButton.clicked.connect(self.Test_btn_clicked)
        
        self.tableWidget.setColumnCount(MAX_COL_SIZE)
        self.tableWidget.setRowCount(MAX_ROW_SIZE)
        
        self.tableWidget.setHorizontalHeaderLabels(['A','B','C','D','E']) 
        
        self.file_name = [] #Excel file name
        self.wb = [] #work book
        self.ws = [] #work sheet
       
        
    def Load_btn_clicked(self):
        file_dlg = QFileDialog()
        self.file_name = file_dlg.getOpenFileName(self, 'Open file', os.getcwd(), "Excel files (*.xlsx)")
        print (self.file_name)
        self.wb = openpyxl.load_workbook(self.file_name[0]) # data_only=False is deafault value
                                                            # data_only=True read just result of function
        self.wb.guess_types = True #type guess enable
        self.ws = self.wb.active
        
        global MAX_ROW_SIZE
        global MAX_COL_SIZE

        MAX_ROW_SIZE = self.ws.max_row
        MAX_COL_SIZE = self.ws.max_column
        
        #change size from sheet size
        self.tableWidget.setColumnCount(MAX_COL_SIZE)
        self.tableWidget.setRowCount(MAX_ROW_SIZE)
        #######
        ## TODO : From MAX SIZE need to update tableWidget.setHorizontalHeaderLabels ( List of label )
        list = Exlcol(MAX_COL_SIZE)
        self.tableWidget.setHorizontalHeaderLabels ( list.list_up() )
        #######
        
        for row in range(1, MAX_ROW_SIZE+1):
            for col in range(1, MAX_COL_SIZE+1):
                cell_txt = self.ws.cell(column = col , row = row).value
                #print(cell_txt) #for debug
                if cell_txt is not None:
                    if(is_number(cell_txt)): #nubmer
                        self.tableWidget.setItem( row-1,col-1, QTableWidgetItem( str(cell_txt) ) )
                    else:
                        self.tableWidget.setItem( row-1,col-1, QTableWidgetItem( cell_txt ) )
                
        
    def Save_btn_clicked(self):
        #print(self.tableWidget.itemAt(1,1).text())
        cell_txt = self.tableWidget.currentItem().text()
        global MAX_ROW_SIZE
        global MAX_COL_SIZE
        
        for row in range(1, MAX_ROW_SIZE+1):
            for col in range(1, MAX_COL_SIZE+1):
                item = self.tableWidget.item(row-1,col-1)
                if item is not None: #QtableWidget item Exist
                    cell_txt = self.tableWidget.item(row-1,col-1).text()
                    if(is_number(cell_txt)): #nubmer
                        self.ws.cell(column = col , row = row, value = cell_txt).number_format
                    else: #text
                        self.ws.cell(column = col , row = row, value = cell_txt)
                else: #Need to make QtableWidget item 
                    self.ws.cell(column = col , row = row, value = "")
        
        self.wb.save(self.file_name[0])
        QMessageBox.about(self, "message", "Save Success!!!")
        
    def Test_btn_clicked(self):
        #cell_txt = self.tableWidget.currentItem().text()
        cur_col = self.tableWidget.currentColumn()
        cur_row = self.tableWidget.currentRow()
        print("%s %s" %(cur_row,cur_col))
        
        #QTableWidget Read Test OK
        # item = self.tableWidget.item(cur_row,cur_col) 
        # if item is not None: #QtableWidget item Exist
            # item.setText("TEST")
        # else: #Need to make QtableWidget item 
            # self.tableWidget.setItem(cur_row,cur_col, QTableWidgetItem("TEST"))
            #cell_txt = self.tableWidget.item(cur_row,cur_col).text()
            #QMessageBox.about(self, "message", cell_txt)
            
        #QTableWidget Write Test
        item = self.tableWidget.item(cur_row,cur_col) 
        if item is not None: #QtableWidget item Exist
            item.setText("TEST")
        else: #Need to make QtableWidget item 
            self.tableWidget.setItem(cur_row,cur_col, QTableWidgetItem("TEST"))
        
if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = MyWindow()
    myWindow.show()
    app.exec_()



